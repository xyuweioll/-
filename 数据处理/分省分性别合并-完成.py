import pandas as pd
import os.path
import win32com.client as win32
import openpyxl
from openpyxl.utils import get_column_letter
import re

# =====================================================================================================================
# 根目录
a = input("是否保留原格式文件？不保留请输入 1/保留 2\n--->>")
b = int(a)
rootdir = u'C:\\Users\\xianyu\\Desktop\\河北分年龄性别 - 副本'  # 存放文件的根目录
# 三个参数：父目录；所有文件夹名（不含路径）；所有文件名
for parent, dirnames, filenames in os.walk(rootdir):
    for fn in filenames:
        if os.path.splitext(fn)[1] == '.xls':  # 判断文件扩展名是不是为.xlsx  os.path.splitext(“文件路径”)
            print(fn)
            filedir = os.path.join(parent, fn)  # os.path.join是路径拼接函数，
            print(filedir)
            # excel = win32.gencache.EnsureDispatch('Excel.Application')   # 若报错则用下一条语句
            excel = win32.DispatchEx('Excel.Application')  # 若上一条报错则使用本条语句
            wb = excel.Workbooks.Open(filedir)  # 打开文件
            # xlsx: FileFormat=51
            # xls:  FileFormat=56,
            # 后缀名的大小写不通配，需按实际修改：xls，或XLS
            wb.SaveAs(filedir.replace('xls', 'xlsx'), FileFormat=51)  # 我这里原文件是小写的xls
            wb.Close()
            excel.Application.Quit()
    if b == 1:
        for fn in filenames:
            if os.path.splitext(fn)[1] == '.xls':  # 判断文件扩展名是不是为.xlsx  os.path.splitext(“文件路径”)
                filedir = os.path.join(parent, fn)  # os.path.join是路径拼接函数，
                os.remove(filedir)
                print(filedir)
print('格式转换完成！')
# ====================================================================================================
for parent, dirnames, filenames in os.walk(rootdir):
    for i in filenames:
        if os.path.splitext(i)[1] == '.xlsx':  # 判断文件扩展名是不是为.xlsx  os.path.splitext(“文件路径”)
            filedir = os.path.join(parent, i)  # os.path.join是路径拼接函数
            print(filedir)
            wb = openpyxl.load_workbook(filedir)
            ws_name = wb.sheetnames  # 获取工作簿中的所有表名
            ws = wb[ws_name[0]]  # 获取第一张表
            ws.delete_rows(1)  # 删除第一行
            wb.save(filedir)
    #  ====================================================================================================================
    # # ----
    # 这一部分实现单元格拆分
    for i in filenames:
        if os.path.splitext(i)[1] == '.xlsx':  # 判断文件扩展名是不是为.xlsx  os.path.splitext(“文件路径”)
            filedir = os.path.join(parent, i)
            data = pd.read_excel(filedir)
            data.fillna(method='ffill', inplace=True)
            data.to_excel(filedir, index=False)
            print(f'{i}单元格拆分已完成！')
    #  ====================================================================================================================
    # 单元格填充
    for i in filenames:
        filedir = os.path.join(parent, i)
        if os.path.splitext(i)[1] == '.xlsx':  # 判断文件扩展名是不是为.xlsx  os.path.splitext(“文件路径”)
            wb = openpyxl.load_workbook(filedir)
            ws_name = wb.sheetnames  # 获取工作簿中的所有表名
            ws = wb[ws_name[0]]  # 获取第一张表
            ws['A1'] = '年龄分组'  # 对单元格A1进行操作，即重命名
            ws.delete_cols(2, 3)  # 删除从第2列开始的3列内容
            ws.delete_rows(2, 2)  # 删除从第2行开始的两行
            r_n = ws.max_row  # 获取行数
            c_n = ws.max_column  # 获取列数
            print(f'行数为{r_n}！, 列数为{c_n}！')
            for j in range(2, c_n + 1, 6):
                ws.cell(row=1, column=j + 1, value=ws.cell(row=1, column=j).value + '女性发病数')
                ws.cell(row=1, column=j + 2, value=ws.cell(row=1, column=j).value + '总计发病数')
                ws.cell(row=1, column=j + 3, value=ws.cell(row=1, column=j).value + '男性死亡数')
                ws.cell(row=1, column=j + 4, value=ws.cell(row=1, column=j).value + '女性死亡数')
                ws.cell(row=1, column=j + 5, value=ws.cell(row=1, column=j).value + '总计死亡数')
                ws.cell(row=1, column=j, value=ws.cell(row=1, column=j).value + '男性发病数')
            print(f'{i}填充已完成！')

            # =====================================================================================================================
            # 插入省份及时间标签列
            Time_list = ['时间']
            Time_lab = [str(os.path.splitext(i)[0][-6:-2] + '/' + os.path.splitext(i)[0][-2:])] * (
                        r_n - 1)  # 根据文件名获取时间标签，并生成列表,r_n为行数
            Time_list.extend(Time_lab)
            Pro_list = ['省份']
            Pro_lab = [os.path.splitext(i)[0][:-6]] * (r_n - 1)  # 根据文件名获取地方标签,并生成列表,r_n为行数
            Pro_list.extend(Pro_lab)
            ws.insert_cols(1)  # 在第一列前插入一列省份列
            for ii in range(r_n):  # 填充值
                ws.cell(ii + 1, 1, Pro_list[ii])
            ws.insert_cols(2)  # 在第二列前插入一列时间列
            for jj in range(r_n):  # 填充值
                ws.cell(jj + 1, 2, Time_list[jj])  # 第一个是行数，第二个值是列数
            print(Time_list, Pro_list)
            # ==================================================================================================================
            # 设置自适应列宽
            # 设置一个字典用于保存列宽数据
            dims = {}
            # 遍历表格数据，获取自适应列宽数据
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                        # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                        # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                        cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                        dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
            for col, value in dims.items():
                # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
                ws.column_dimensions[get_column_letter(col)].width = value + 3
            wb.save(filedir)
# ====================================================================================================================
#  纵向合并
concat_list = []
for parent, dirnames, filenames in os.walk(rootdir):
    for i in filenames:
        if os.path.splitext(i)[1] == '.xlsx':  # 判断文件扩展名是不是为.xlsx  os.path.splitext(“文件路径”)
            filedir = os.path.join(parent, i)
            concat_list.append(pd.read_excel(filedir, sheet_name=0))
            print(f'文件{i}读取完成！')
out_file = concat_list[0]
for i in range(1, len(concat_list)):
    print(f'第{i+1}个文件合并中!')
    out_file = pd.concat([out_file, concat_list[i]], ignore_index=True)
out_path = r'C:\Users\xianyu\Desktop\output_file.csv'  # 根据实际情况更改
out_file.to_csv(out_path, index=False)
# ===========================================================================================================
# 调整输出文件的列宽
# wb = openpyxl.load_workbook(out_path)
# ws_name = wb.sheetnames  # 获取工作簿中的所有表名
# ws = wb[ws_name[0]]  # 获取第一张表
# # 设置自适应列宽
# # 设置一个字典用于保存列宽数据
# dims = {}
# # 遍历表格数据，获取自适应列宽数据
# for row in ws.rows:
#     for cell in row:
#         if cell.value:
#             # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
#             # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
#             # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
#             cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
#             dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
# for col, value in dims.items():
#     # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+4是用来调整最终效果的
#     ws.column_dimensions[get_column_letter(col)].width = value + 4
# wb.save(out_path)
print(f'文件已输出，请至目录{out_path}查看！')
