import pandas as pd
import os
import numpy as np           # G:\\回收站\\东北Satscan\\12-19东北地区处理前 - 副本.xlsx     Sheet1
from pandas import DataFrame
from openpyxl import Workbook
input('提示信息：本程序默认将第一列作为地名或地名标识符，第二列作为行政编码，请将数据调整至相应格式\n若已调整按任意键继续===>>>')
file_path = str(input('请输入文件路径，如:C:\\\\Users\\\\xianyu\\\\Desktop\\\\东北Satscan\\\\12-16东北地区处理前.xlsx\n===>>>'))
print(file_path)
if os.path.splitext(file_path)[1] =='.xlsx' or os.path.splitext(file_path)[1] =='.xls':
    sheet_name = input('请输入表名\n===>>>')
    data = pd.read_excel(file_path, sheet_name=sheet_name)
else:
    data = pd.read_csv(file_path)
start_col = int(input('请输入病例数据从第几列开始===>>>'))
year_start = int(input('请输入开始年份===>>>'))
year_num = int(input('请输入数据包含的年份===>>>'))
# 读取数据及获取参数
# ============================================================================================
label_name = []   # 中间列表，只起到临时存放作用
a = np.array(data.iloc[:, 0:1].values).tolist()
[label_name.extend(i) for i in a]  # 此时label_name是各地方的名字
a = label_name.copy()  # 一定要用copy的方式赋值，直接赋值只会将地址赋值，a的值会随着label_name改变，获取地方名列表，及第一列
label_name.clear()  # 清空label_name供下一次使用
b = np.array(data.iloc[:, 1:2].values).tolist()  # 获取地方行政编码列表,即第2列
[label_name.extend(i) for i in b]
b = label_name.copy()
label_name.clear()
# a,b两个列表已经成功获取
# ============================================================================================
# 获取病例数列表
c = []  # 中间用于临时存放的列表
for i in range(len(a)):
    xuan_hang = [i]  # 获取特定行的数据，i为行数
    d = data.iloc[xuan_hang, start_col-1:start_col+year_num*12-1]  # print(c)#获取特定行中的一些列,但此时d还是一个DataFrame数据类型
    d_array = np.array(d)  # 将DataFrame转换成array
    d_list = d_array.tolist()  # 将array 转换成列表
    c.extend(d_list)  # 将列表合并,此时列表中的元素还是一个列表，每一个列表是一个市的病例数
d = []
for i in c:
    d.extend(i)  # 实现将列表里的列表合并成一个大列表，元素是单个数值，此时d为病例数的列表
print(f'd has {len(d)} factors!')
# ===============================================================================================
total_row = data.shape[0]     # 获取列表总行数，行数即城市个数
date_list = []  #
year_list = [i for i in range(year_start, year_start+year_num)]
month = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']
for k in year_list:
    for l in month:
        m = str(k) + '/' + l
        date_list.append(m)  # 列表date_list 是日期列表
date_list = date_list*total_row   # total_row是城市个数
print(f'data_list has {len(date_list)} factors!')
# ====================================================================================================
# 人口模块生成，默认人口是紧接着病例数后进行排列的
cc = []  # 用于临时存放的列表
for i in range(len(a)):
    xuan_hang = [i]  # 获取特定行的数据，i为行数
    pop = data.iloc[xuan_hang, start_col+year_num*12-1:start_col+year_num*12-1+year_num]  # print(c)#获取特定行中的一些列,但此时d还是一个DataFrame数据类型
    p_array = np.array(pop)   # 将DataFrame转换成array
    p_list = p_array.tolist()  # 将array 转换成列表
    cc.extend(p_list)  # 将列表合并,此时列表中的元素还是一个列表，每一个列表是一个市的病例数
pop_list = []
for i in cc:
    pop_list.extend(i)  # 实现将列表里的列表合并成一个大列表，元素是单个数值，此时pop_list为人口数的列表
pop_list = [val for val in pop_list for i in range(12)]     #  每个元素重复12遍,区别于 pop_list *= 12
print(f'pop_list has {len(pop_list)} factors!')
# ====================================================================================================
# 生成经度，其默认是紧接着人口数排列
ccc = []  # 用于临时存放的列表
for i in range(len(a)):
    xuan_hang = [i]  # 获取特定行的数据，i为行数
    lon = data.iloc[xuan_hang, start_col+year_num*12-1+year_num:start_col+year_num*12+year_num]  # print(c)#获取特定行中的一些列,但此时d还是一个DataFrame数据类型
    l_array = np.array(lon)   # 将DataFrame转换成array
    l_list = l_array.tolist()  # 将array 转换成列表
    ccc.extend(l_list)  # 将列表合并,此时列表中的元素还是一个列表，每一个列表是一个市的病例数
lon_list = []
for i in ccc:
    lon_list.extend(i)  # 实现将列表里的列表合并成一个大列表，元素是单个数值，此时pop_list为人口数的列表
lon_list = [val for val in lon_list for i in range(12*year_num)]     #  每个元素重复12遍,区别于 pop_list *= 12
print(f'lon_list has {len(lon_list)} factors!')
# ====================================================================================================
# 生成纬度，其默认是紧接着人口数排列
cccc = []  # 用于临时存放的列表
for i in range(len(a)):
    xuan_hang = [i]  # 获取特定行的数据，i为行数
    lat = data.iloc[xuan_hang, start_col+year_num*12+year_num:start_col+year_num*12+year_num+1]  # print(c)#获取特定行中的一些列,但此时d还是一个DataFrame数据类型
    ll_array = np.array(lat)   # 将DataFrame转换成array
    ll_list = ll_array.tolist()  # 将array 转换成列表
    cccc.extend(ll_list)  # 将列表合并,此时列表中的元素还是一个列表，每一个列表是一个市的病例数
lat_list = []
for i in cccc:
    lat_list.extend(i)  # 实现将列表里的列表合并成一个大列表，元素是单个数值，此时pop_list为人口数的列表
lat_list = [val for val in lat_list for i in range(12*year_num)]     #  每个元素重复12遍,区别于 pop_list *= 12
print(f'lat_list has {len(lat_list)} factors!')
# ====================================================================================================
# 创建excel表格
fileName = "SaTScan1.xlsx"
wb = Workbook()  # 创建一个Workbook，表名默认为Sheet
ws = wb.active  # 激活一张表格,因为此时只有一张表格，默认为Sheet,激活表格后若不选表格则默认操作激活表格
# ws = wb["Sheet"]  #选中某一张特定表格
print("默认创建的sheet名", ws.title)
ws.append(['NAME', 'CITY_CODE', 'CASES', 'DATE', 'POP', 'LON', 'LAT'])
f = 0
for i in range(len(a)):
    for j in range(year_num*12):    # 月数
        ws.append([a[i], b[i], d[f], date_list[f], pop_list[f], lon_list[f], lat_list[f]])
        f += 1
wb.save("C:\\Users\\xianyu\\Desktop\\SaTScan1.xlsx")








# 默认会创建一个名为“Sheet”的sheet
# ws = wb.active
# #修改sheet的名称
# ws.title = "Sheet1"
# #创建新sheet,可以控制位置，都表示是在这个位置之前插入
# ws1 = wb.create_sheet("Mysheet1")#默认在最后位置插入
# ws2 = wb.create_sheet("Mysheet2", 0)#0是指第一张表，在第一张表前面插入一张表
# ws3 = wb.create_sheet("Mysheet3", -1)#表示在最后一张表的前面插入一张表
#
# #遍历所有的sheet
# for sheetname in wb.sheetnames:
#     print("遍历Sheet", sheetname)
#
# #选中sheet
# ws3 = wb["Mysheet3"] #选中一张表格
# print("选中sheet name:", ws3.title)
# #修改active sheet
# wb.active = wb["Mysheet3"]
# ws3 = wb.active
# print("当前active sheet:", ws3.title)
#
# #单元格层面
# #单元格赋值
# ws3["A1"] = 1
# cell = ws3["A1"]
# print("A1的值", cell.value)

# fname = "C:\\myProjects\\daily\\data\\20171229\\标准化文档-20171226\\全量数据过滤表清单.xls"
# wb = excel.Workbooks.Open(fname) #打开文件
