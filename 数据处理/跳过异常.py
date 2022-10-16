import re    # 这个包的作用是根据多个指定字符对字符串进行分割
import pandas as pd
import os
import openpyxl
pname = input("请输入你要查询的省份名\n--->>>")
dname = input("请输入你要查询的病种名\n--->>>")
a = input("请输入你要查询的日期范围\n（输入示例：如2001年6月至2020年4月：2001,06:2020,04）"
          "注意！间隔符为英文形式\n--->>>")
b = re.split(r'[,:]\s*', a)  # 调用re包的split方法指定根据字符“,”和“:”进行分隔，返回列表
# 此时b的形式为:['2020', '02', '2020', '09']
c = []
for i in b:
    c.append(int(i))
# print(c)  #这个循环的作用是将b中的每一个字符串转换成整型数据病存放到列表c中
d = (c[2]-c[0])*12+(c[3]-c[1])+1  # d的数值是两个日期之间包含的月数
# print(d)

# 这段循环的作用是生成文件名后面的时间标签，并存放在列表e中,其元素格式如201001
e = []
if c[0] == c[2]:
    for i in range(c[1], c[3]+1):
        if i < 10:
            e.append(str(c[0]) + '0' + str(i))
        else:
            e.append(str(c[0]) + str(i))
else:
    for i in range(c[1], 13):
        if i < 10:
            e.append(str(c[0])+'0'+str(i))
        else:
            e.append(str(c[0])+str(i))
for i in range(c[0]+1, c[2]):
    for j in ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']:
        e.append(str(i)+j)
if c[0] != c[2]:
    for i in range(1, c[3]+1):
        if i < 10:
            e.append(str(c[2])+'0'+str(i))
        else:
            e.append(str(c[2])+str(i))
print('程序段1执行完毕')

# print(e)#此时e是完整的时间标签列表
# 这个循环的作用是生成文件名，即将地方名和日期标签进行连接生成一个列表，这个列表里是表名
f = []
g = []
exception = []   # 该列表用于存储月份表中该病不存在的表名
for i in e:
    f.append(pname+i)  # 此时列表f已经是需要查询的文件名,不带后缀
    g.append(pname+i+'.xlsx')  # 此时文件名列表里文件的后缀是xlsx   此时列表f已经是需要查询的文件名
# print(f)

os.chdir('C:\\Users\\xianyu\\Desktop\\T')  # 放处理文件的路径，根据实际情况更改
h = []
p_m = []  # p_m 列表的作用是写每个月每张表时使用
feature_name0 = ['地区', 'CITY_CODE', dname, dname+'death']
for i in range(len(f)):
    try:
        m = pd.read_excel(g[i])    # 读入文件/这个是在写每张表时用:p_m
        p_m.append(m[feature_name0])
    except Exception as oo:
        exception.append(f[i])
    continue
# -----------------------------
m = pd.read_excel(g[0])
# print(type(m))  #m此时是DataFrame数据格式
try:
    h.append(m[feature_name0])  # 连接用的第一张表
except []:
    print(f'表{f[0]}中没有{dname}')

# print(h)
feature_name1 = ['CITY_CODE', dname, dname+'death']      # 若想在合并表里保留名字，在前面加'地区'

for i in range(1, len(f)):             # 这个循环生成的结果是连接时使用
    try:
        m = pd.read_excel(g[i])
        h.append(m[feature_name1])   # 加 try  except 的是因为有的病在部分月份里没有，故跳出一次循环，并把表名列表f中对应的
    except Exception as ooo:         # 表名删除
        print(f'表{f[i]}中没有{dname}')  # 打印提示
        pass

    continue
print('代码段3执行完毕！')
# --------------------------
# 这段程序将刚在出现异常的月份表在f列表里删除
for i in exception:
    print(f'{i}已删除！')
    f.remove(i)
# --------------------------
# 将每个月的数据分别写成一个sheet
writer = pd.ExcelWriter('C:\\Users\\xianyu\\Desktop\\输出数据.xlsx')
for i in range(len(h)):
    p_m[i].to_excel(writer, sheet_name=f[i], index=False)
writer.save()

# -----------------------
# 是合并这个模块的作用，以’CITY_CODE‘匹配进行横向追加
for i in range(1, len(h)):
    h[0] = pd.merge(h[0], h[i], left_on='CITY_CODE', right_on='CITY_CODE', how='left')  # 左连接 how = 'left'
# print(h[0])                                                                            # 右连接 how ='right'
# -------------                                                                          # 全连接 how = 'outer'
# 写出文件(合并表)
h[0].to_excel(writer, sheet_name='合并表', index=False)
writer.save()
writer.close()
print('代码段4执行完毕！')
# ------------------------------
# 改表头为日期
wb = openpyxl.load_workbook('C:\\Users\\xianyu\\Desktop\\输出数据.xlsx')
ws = wb['合并表']
r_n = ws.max_row  # 获取行数
c_n = ws.max_column  # 获取列数
z = 0
for i in range(3, c_n, 2):
    ws.cell(row=1, column=i, value=f[z])
    ws.cell(row=1, column=i+1, value=f[z])
    z += 1
wb.save('C:\\Users\\xianyu\\Desktop\\输出数据.xlsx')