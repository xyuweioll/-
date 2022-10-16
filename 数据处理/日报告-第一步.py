import pandas as pd
import os.path
import win32com.client as win32
import openpyxl
a = input("是否保留原格式文件？不保留请输入 1/保留 2\n--->>")
b = int(a)
rootdir = u'C:\\Users\\xianyu\\Desktop\\调试数据'  # 存放文件的根目录
for parent, dirnames, filenames in os.walk(rootdir):
    for fn in filenames:
        filedir = os.path.join(parent, fn)  # os.path.join是路径拼接函数，
        print(filedir)

        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(filedir)  # 打开文件
        # xlsx: FileFormat=51
        # xls:  FileFormat=56,
        # 后缀名的大小写不通配，需按实际修改：xls，或XLS
        wb.SaveAs(filedir.replace('xls', 'xlsx'), FileFormat=51)  # 我这里原文件是小写的xls
        wb.Close()
        excel.Application.Quit()
    # 这一段代码实现xls到xlsx的转换
# -------------------------------------------
    # 这一段代码的作用是删除原文件夹.xls文件
    for fn in filenames:
        filedir = os.path.join(parent, fn)  # os.path.join是路径拼接函数，
        os.remove(filedir)
        print(filedir)
# -------------------------------------------------------------------------------
#实现删除第一行
a = []
for parent, dirnames, filenames in os.walk(rootdir):
    for i in filenames:
        b = len(i)-5
        a.append(i[:b])  # 这个循环的作用是去除后缀.xlsx（即去除后五个字符）
    print(a)
    for i in range(len(filenames)):
        filedir = os.path.join(parent, filenames[i])  # os.path.join是路径拼接函数，
        print(filedir)
        wb = openpyxl.load_workbook(filedir)
        ws = wb[a[i]]
        ws.delete_rows(1)  # 删除第一行
        wb.save(filedir)
# # ----
    # 这一部分实现单元格拆分
    for i in filenames:
        filedir = os.path.join(parent, i)
        data = pd.read_excel(filedir)
        data.fillna(method='ffill', inplace=True)
        data.to_excel(filedir, index=False)
# # -----
    # 这一部分代码的作用是填充值
    for i in filenames:
        filedir = os.path.join(parent, i)
        wb = openpyxl.load_workbook(filedir)
        ws = wb['Sheet1']
        ws['A1'] = '时间'  # 对单元格A1进行操作
        r_n = ws.max_row  # 获取行数
        c_n = ws.max_column  # 获取列数
        print(f'{i}的行数：{r_n}  列数： {c_n}')
        for j in range(4, c_n + 1, 2):
            ws.cell(row=1, column=j, value=ws.cell(row=1, column=j - 1).value+'death')
        wb.save(filedir)
# # # ------------------------------------------------------------------------------------------------