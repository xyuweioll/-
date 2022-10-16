import pandas as pd
import os.path
import win32com.client as win32
import openpyxl
rootdir = u'C:\\Users\\xianyu\\Desktop\\test'  # 存放文件的根目录
for parent, dirnames, filenames in os.walk(rootdir):
    for fn in filenames:
        filedir = os.path.join(parent, fn)  # os.path.join是路径拼接函数，
        print(filedir)

        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(filedir)  # 打开文件
        # xlsx: FileFormat=51
        # xls:  FileFormat=56,
        # 后缀名的大小写不通配，需按实际修改：xls，或XLS
        wb.SaveAs(filedir.replace('xls', 'xlsx'), FileFormat=51)  # 我这里原文件是小写的xls
        wb.Close()
        excel.Application.Quit()
    # 这一段代码实现xls到xlsx的转换
# -------------------------------------------
    # 这一段代码的作用是删除原文件夹.xls文件

    for fn in filenames:
        filedir = os.path.join(parent, fn)  # os.path.join是路径拼接函数，
        os.remove(filedir)
        print(filedir)
# -------------------------------------------------------------------------------
#
a = []
for parent, dirnames, filenames in os.walk(rootdir):
    for i in filenames:
        b = len(i)-5
        a.append(i[:b])  # 这个循环的作用是去除后缀.xlsx（即去除后五个字符）
    print(a)
    for i in range(len(filenames)):
        filedir = os.path.join(parent, filenames[i])  # os.path.join是路径拼接函数，
        print(filedir)
        wb = openpyxl.load_workbook(filedir)
        ws = wb[a[i]]
        ws.delete_rows(1)  # 删除第一行
        wb.save(filedir)
# # ----
    # 这一部分实现单元格拆分
    for i in filenames:
        filedir = os.path.join(parent, i)
        data = pd.read_excel(filedir)
        data.fillna(method='ffill', inplace=True)
        data.to_excel(filedir, index=False)
# # -----
    # 这一部分代码的作用是填充值
    for i in filenames:
        filedir = os.path.join(parent, i)
        wb = openpyxl.load_workbook(filedir)
        ws = wb['Sheet1']
        ws['A1'] = '地区'  # 对单元格A1进行操作
        ws['B1'] = 'CITY_CODE'  # 对单元格B1进行操作
        ws['B2'] = 1
        r_n = ws.max_row  # 获取行数
        c_n = ws.max_column  # 获取列数
        print(r_n, c_n)
        for j in range(5, c_n + 1, 2):
            ws.cell(row=1, column=j, value=ws.cell(row=1, column=j - 1).value+'death')
        wb.save(filedir)
# ------------------------------------------------------------------------------------------------
year = []
p_name = []
for parent, dirnames, filenames in os.walk(rootdir):
    for i in filenames:
        b = len(i) - 5
        year.append(i[:4])  # 获取年份
        p_name.append(i[4:b])  # 获取省份名
year = list(set(year))
p_name = list(set(p_name))
print(year)
print(p_name)

#  -------------------------------------------------------------------------------------------------
os.chdir('C:\\Users\\xianyu\\Desktop\\test')   # 存放目录的文件
for y in year:
    file_list = []
    for p in p_name:
        try:
            file_name = y+p+'.xlsx'
            re_file = pd.read_excel(file_name)
            file_list.append(re_file)
        except:
            print(f'{parent}中没有{file_name}文件，请检查！')
    print(f'{y}年有{len(file_list)}个文件！')    # 用于检查每年各个省份的文件是否齐全
# ----------------------------------------------------------------------------------
# 分年份对各省份的文件进行连接

    zong = file_list[0]
    for i in range(1, len(file_list)):
        zong = pd.concat([zong, file_list[i]], ignore_index=True)
    output_name = y+'合并文件'+'.xlsx'
    zong.to_excel(f'C:\\Users\\xianyu\\Desktop\\{output_name}', index=False)
print('程序已执行完毕，在桌面查看输出文件')



