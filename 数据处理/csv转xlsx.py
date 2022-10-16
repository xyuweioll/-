# import pandas as pd
# import os
# os.chdir('C:\\Users\\xianyu\\Desktop') #确定操作路径,用os确定路径时一定要用\\，不可以用\
# # 用pandas包实现csv转excel
# a = "上海201002.csv"
# csv = pd.read_csv(a)
# csv.to_excel('上海201002.xlsx', index=False)
#
# #用pandas包实现excel转csv
# # excel = pd.read_excel("上海201002.xlsx")
# # excel.to_csv('2.csv', encoding='utf-8', index=False)  # index = Flase 的目的是不需要所有列

#---------------------------------------------------------------------------------
# 拆分合并单元格模块（对纵向拆分的单元格进行合并）
# import openpyxl
# import os
# os.chdir('C:\\Users\\xianyu\\Desktop')
# workbook = openpyxl.load_workbook('上海201002.xlsx')  # 加载已经存在的excel
# name_list = workbook.sheetnames      # 获取表名
# worksheet = workbook[name_list[0]]  #打开第一张表
# m_list = worksheet.merged_cells
# if m_list:             # 判断m_list是不是空，若不是空再进行下一步，# 合并单元格的位置信息，可迭代对象excel坐标信息# print(m_list)
#     cr = []            #（单个是一个'openpyxl.worksheet.cell_range.CellRange'对象），print后就是
#     for m_area in m_list:
#         # 合并单元格的起始行坐标、终止行坐标。。。。，
#         r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
#         # 纵向合并单元格的位置信息提取出
#         if r2 - r1 > 0:
#             cr.append((r1, r2, c1, c2))
#     # 这里注意需要把合并单元格的信息提取出再拆分
#     for r in cr:
#         worksheet.unmerge_cells(start_row=r[0], end_row=r[1],
#                                 start_column=r[2], end_column=r[3])
#         for i in range(r[1] - r[0] + 1):
#             for j in range(r[3] - r[2] + 1):
#                 worksheet.cell(row=r[0] + i, column=r[2] + j, value=worksheet.cell(r[0], r[2]).value)
#     workbook.save('上海201001.xlsx')



#  ----------------------------------------------------------------------------------------------------------
# 拆分单元格合并(横向及纵向合并并纵向填充值)
import pandas as pd
import os
import openpyxl
os.chdir('C:\\Users\\xianyu\\Desktop')
data = pd.read_excel(r'上海201002.xlsx')
data.fillna(method='ffill', inplace=True)
data.to_excel(r'上海201002.xlsx', index=False)
#这一部分实现单元格拆分
#-----------------------------------------------
#这一部分实现值填充
wb = openpyxl.load_workbook('上海201002.xlsx')  # 加载已经存在的excel
ws = wb['Sheet1']
r_n = ws.max_row  # 获取行数
c_n = ws.max_column  # 获取列数
print(r_n, c_n)
for i in range(5, c_n+1, 2):
    ws.cell(row=1, column=i, value=ws.cell(row=1, column=i-1).value)
wb.save('上海201002.xlsx')

# ------------------------------------------------------------------------------------------------------------
# # 拆分多行多列合并的单元格
# import openpyxl
# import os
# os.chdir('C:\\Users\\xianyu\\Desktop')
# workbook = openpyxl.load_workbook('上海201002.xlsx')  # 加载已经存在的excel
# name_list = workbook.sheetnames      # 获取表名
# worksheet = workbook[name_list[0]]  #打开第一张表
# m_list = worksheet.merged_cells
# if m_list:             # 判断m_list是不是空，若不是空再进行下一步，# 合并单元格的位置信息，可迭代对象excel坐标信息# print(m_list)
#     cr = []            #（单个是一个'openpyxl.worksheet.cell_range.CellRange'对象），print后就是
#     for m_area in m_list:
#         # 合并单元格的起始行坐标、终止行坐标。。。。，
#         r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
#         # 纵向合并单元格的位置信息提取出
#         if r2 - r1 > 0:
#             cr.append((r1, r2, c1, c2))
#     # 这里注意需要把合并单元格的信息提取出再拆分
#     for r in cr:
#         for i in range(r[1] - r[0] + 1):
#             for j in range(r[3] - r[2] + 1):
#                 worksheet.cell(row=r[0] + i, column=r[2] + j, value=worksheet.cell(r[0], r[2]).value)
#         # worksheet.unmerge_cells(start_row=r[0], end_row=r[1],
#         #                         start_column=r[2], end_column=r[3])
#         # for i in range(r[1] - r[0] + 1):
#         #     for j in range(r[3] - r[2] + 1):
#         #         worksheet.cell(row=r[0] + i, column=r[2] + j, value=worksheet.cell(r[0], r[2]).value)
#     workbook.save('上海201001.xlsx')
